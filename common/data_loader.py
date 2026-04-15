"""==========
■ データ取得
=========="""

import os
from datetime import date, datetime
from typing import Any, Optional

## gspreadの操作
import gspread
import openpyxl
import pandas as pd
from google.oauth2.service_account import Credentials

from common.log_handler import log_error, log_info

# rootなど
from config.settings import (
    DATA_SOURCE,
    KEY_FILE_NAME,
    PROJECT_ROOT,
    SHEET_NAME,
    SPREADSHEET_NAME,
)


def find_key_path(KEY_FILE_NAME: str):
    """Jsonキーファイルのパスをプロジェクト全体から探索して返します。

    Args:
        filename (_type:str_):
    """
    try:
        # 探索パスリスト
        search_paths = [
            PROJECT_ROOT / "config" / KEY_FILE_NAME,
            PROJECT_ROOT / KEY_FILE_NAME,
        ]

        # Jsonキーのパスを探索リストをもとに探索し、見つかったらパスを文字列で返す
        for search_path in search_paths:
            if search_path.exists():
                return str(search_path)
        log_error(f"キーファイル{KEY_FILE_NAME}が見つかりませんでした")
        return None
    except Exception:
        log_error("キーファイル探索中にエラーが発生しました")
        return None


def load_data(file_path: Optional[str] = None) -> list[list[Any]]:
    """_summary_設定（DATA_SAUCE）に基づいて、ExcelまたはGoogleシートからデータを読み込む

    Args:
        file_path (_type_, optional): _description_. Defaults to None.

    Returns:
        List: 行データのリスト（ヘッダを除くん次元配列
        ※Googleシートの場合もExcel互換の形式（日付obj）に変換して返す
    """
    if DATA_SOURCE == "GOOGLE":
        log_info(
            f"☁️ [Data Load] Google Sheets ('{SPREADSHEET_NAME}') からデータを読み込みます..."
        )
        return _load_from_gspread()
    else:
        log_info(f"📂 [Data Load] Excel ('{file_path}') からデータを読み込みます...")
        return _load_from_excel(file_path)


def _load_from_excel(file_path):
    """_summary_:既存のExcel読み込みロジック"""
    if not file_path or not os.path.exists(file_path):
        raise FileNotFoundError(f"Excelファイルが見つかりません： {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.worksheets[0]

    # values_only=true で値のリストとして取得（ヘッダを除く二行目から）
    data = list(ws.iter_rows(min_row=2, values_only=True))
    return data


def _load_from_gspread():
    """_summary_:Google Sheets読み込みロジック（Excel互換形式に変換）"""
    key_path = find_key_path(KEY_FILE_NAME)
    if not key_path:
        raise FileNotFoundError(
            f"サービスアカウントキー({KEY_FILE_NAME})が見つかりませんでした"
        )

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(key_path, scopes=scopes)
    client = gspread.authorize(creds)

    workbook = client.open(SPREADSHEET_NAME)
    try:
        worksheet = workbook.worksheet(SHEET_NAME)
    except gspread.WorksheetNotFound:
        worksheet = workbook.get_worksheet(0)
        log_error("ワークシートが存在しません")

    # 全データを文字列リストとして取得(get_all_values)
    raw_data = worksheet.get_all_values()

    if len(raw_data) < 2:
        return []

    # ヘッダを除外
    rows = raw_data[1:]

    # excelとの互換性確保のための型変換処理
    # google sheetsはすべて文字列で買えるため、日付列などを日付型に変換する
    converted_rows = []
    # 取得したデータを1行ずつ日付変換する
    for row in rows:
        new_row = list(row)

        # 1. 日付変換（0列目："yyyy/mm/dd" -> datetime object
        if len(new_row) > 0 and new_row[0]:
            try:
                # フォーマットは実際のデータに合わせて調整
                new_row[0] = pd.to_datetime(new_row[0]).to_pydatetime()
            except Exception:
                pass  # 変換できなければそのまま

        # 2. 必要に応じて

        converted_rows.append(new_row)

    return converted_rows


def find_today_row(data_list: list[list[Any]], target_date: date) -> list[Any] | None:
    """_summary_:データリストから指定された日付の行を探して返します

    Args:
        data_list (_type_：List): load_dataで取得した行データのリスト
        target_date (_type_：datetime.date): 探したい日付

    Returns:
        list or tuple: 見つかった行データ。なければNone
    """
    log_info(f"🔍 [Search] {target_date} のデータを探しています...")

    for row in data_list:
        # Excel/Googleシートの空行対策
        if not row or not row[0]:
            continue

        row_date = row[0]
        # row_dateがdatetimeオブジェクトの場合、date型に変換して比較
        if isinstance(row_date, datetime):
            row_date = row_date.date()

        # 文字列のままの場合の救済措置
        elif isinstance(row_date, str):
            try:
                row_date = pd.to_datetime(row_date).date()
            except Exception:
                continue

        if row_date == target_date:
            log_info("🔦 データが見つかりました")
            return row

    log_error("❌ 今日のデータが見つかりませんでした")

    return None
